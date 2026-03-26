"""
================================================================================
pages/admin.py -- Admin Panel
================================================================================
v2 Changes:
    G1 - No sidebar; Back + Logout moved to slim top nav bar
    G2 - All dropdowns non-typeable (st.selectbox everywhere)
    G3 - Bold headings + section separators throughout all tabs
    D1 - Head name / email / phone are now REQUIRED fields
    D2 - Clear section separators in Documents tab
    T1 - Token Usage charts: smaller height, expand toggle, Bar/Line toggle
    T2 - Token Usage date: calendar date picker (st.date_input range)
    H1 - History: remove risk level filter + remove Danger Zone
    E1 - Export: calendar date picker same as Token Usage
    +  - No ">" back arrow; radio changed to "View" / "Upload" / "Manage"
    +  - Header row email/phone validation messages shown cleanly full-width
================================================================================
"""

import os
import glob
import shutil
import json
from datetime import datetime, timedelta, date
from io import BytesIO

import streamlit as st
from dotenv import load_dotenv

from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import LETTER
from reportlab.lib import colors as rl_colors

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import logger
import dept_db

# ==========================================
# CONFIG & PATHS
# ==========================================
load_dotenv()

ADMIN_PASSWORD       = os.getenv("ADMIN_PASSWORD", "admin123")
BASE_DIR             = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONTRACTS_DIR        = os.path.join(BASE_DIR, "MyFiles", "Contracts")
COMPANY_PLAYBOOK_DIR = os.path.join(CONTRACTS_DIR, "company_standard")
CLIENT_CONTRACTS_DIR = os.path.join(CONTRACTS_DIR, "clients")
CHROMA_PERSIST_DIR   = os.path.join(BASE_DIR, "chroma_db")

dept_db.init_db()

st.set_page_config(
    page_title="Admin Panel — Legal Contract Analyzer",
    page_icon="⚙",
    layout="wide",
    # G1: collapse sidebar entirely — nav is in top bar
    initial_sidebar_state="collapsed",
)

# CSS: arrow/menus hidden, top space removed, non-typeable selects, read-only dates, compact buttons
st.markdown("""
    <style>
    /* Fix 1: Remove > sidebar collapse arrow — every known selector */
    [data-testid="stSidebarNav"]                         { display: none !important; }
    [data-testid="stSidebar"]                            { display: none !important; }
    [data-testid="collapsedControl"]                     { display: none !important; }
    [data-testid="stSidebarCollapsedControl"]            { display: none !important; }
    section[data-testid="stSidebar"]                     { display: none !important; }
    button[kind="header"]                                { display: none !important; }
    .st-emotion-cache-czk5ss                             { display: none !important; }
    .st-emotion-cache-1dp5vir                            { display: none !important; }
    .st-emotion-cache-dvne4q                             { display: none !important; }
    [data-testid="baseButton-headerNoPadding"]           { display: none !important; }
    header[data-testid="stHeader"] button                { display: none !important; }
    header[data-testid="stHeader"]                       { display: none !important; }
    [data-testid="stToolbar"]                            { display: none !important; }
    #MainMenu                                            { display: none !important; }
    footer                                               { display: none !important; }

    /* Fix 2: Remove large top empty space */
    [data-testid="stAppViewContainer"]                   { padding-top: 0 !important; }
    [data-testid="block-container"]                      { padding-top: 12px !important; }
    .block-container                                     { padding-top: 12px !important; }

    /* Selectbox non-typeable globally */
    div[data-baseweb="select"] input                     { caret-color: transparent !important;
                                                           pointer-events: none !important; }
    div[data-baseweb="select"] [data-baseweb="input"]    { pointer-events: none !important; }

    /* Date input read-only */
    div[data-baseweb="input"] input[type="text"]         { caret-color: transparent !important;
                                                           user-select: none !important; }
    div[data-baseweb="input"] [data-baseweb="icon"]      { pointer-events: all !important; }

    /* Fix 3: Back to Main App button compact */
    div[data-testid="stButton"] > button[kind="secondary"] {
        font-size: 11px !important;
        padding: 3px 10px !important;
        min-height: 30px !important;
    }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# SESSION STATE
# ==========================================
for key, val in {
    "admin_auth":          False,
    "admin_fail_count":    0,
    "admin_lockout_until": None,
}.items():
    if key not in st.session_state:
        st.session_state[key] = val

# ==========================================
# PASSWORD GATE
# ==========================================
if not st.session_state.admin_auth:
    locked = False
    if st.session_state.admin_lockout_until:
        remaining = (st.session_state.admin_lockout_until - datetime.now()).total_seconds()
        if remaining > 0:
            locked = True
        else:
            st.session_state.admin_lockout_until = None
            st.session_state.admin_fail_count    = 0

    col_l, col_c, col_r = st.columns([1, 1.2, 1])
    with col_c:
        st.markdown("<div style='height:60px'></div>", unsafe_allow_html=True)
        st.markdown(
            '<div style="text-align:center;margin-bottom:24px;">'
            '<div style="font-size:40px;">⚙</div>'
            '<h2 style="margin:8px 0 4px;">Admin Panel</h2>'
            '<p style="color:#6b7280;font-size:14px;margin:0;">'
            'Restricted access. Enter your password to continue.</p>'
            '</div>',
            unsafe_allow_html=True,
        )

        if locked:
            remaining_s = int((st.session_state.admin_lockout_until - datetime.now()).total_seconds())
            st.warning(f"Too many failed attempts. Try again in {remaining_s} seconds.")
        else:
            if st.session_state.admin_fail_count > 0:
                dots = '<div style="display:flex;gap:6px;justify-content:center;margin-bottom:12px;">'
                for i in range(5):
                    color = "#ef4444" if i < st.session_state.admin_fail_count else "#e5e7eb"
                    dots += f'<div style="width:10px;height:10px;border-radius:50%;background:{color};"></div>'
                dots += "</div>"
                st.markdown(dots, unsafe_allow_html=True)

            pwd = st.text_input(
                "Password", type="password",
                label_visibility="collapsed",
                placeholder="Enter admin password...",
            )
            col_a, col_b = st.columns(2)
            with col_a:
                if st.button("Login", type="primary", use_container_width=True):
                    if pwd == ADMIN_PASSWORD:
                        st.session_state.admin_auth       = True
                        st.session_state.admin_fail_count = 0
                        st.rerun()
                    else:
                        st.session_state.admin_fail_count += 1
                        if st.session_state.admin_fail_count >= 5:
                            st.session_state.admin_lockout_until = (
                                datetime.now() + timedelta(seconds=60)
                            )
                            st.rerun()
                        else:
                            rem = 5 - st.session_state.admin_fail_count
                            st.error(f"Incorrect password. {rem} attempt(s) remaining.")
            with col_b:
                if st.button("Cancel", use_container_width=True):
                    st.switch_page("app.py")
    st.stop()


# ==========================================
# TOP NAV BAR
# ==========================================
nav_l, nav_c, nav_r = st.columns([2, 4, 2])
with nav_l:
    # Fix 3: compact back button — not full width
    btn_col, _ = st.columns([1.4, 1])
    with btn_col:
        if st.button("Back to Main App", use_container_width=True):
            st.switch_page("app.py")
with nav_c:
    st.markdown(
        '<div style="text-align:center;padding:2px 0;">' 
        '<span style="font-size:26px;font-weight:800;color:#1E2D5E;">⚙ Admin Panel</span>'
        '<br><span style="font-size:11px;color:#9ca3af;">'
        'Legal Contract Analyzer</span>'
        '</div>',
        unsafe_allow_html=True,
    )
with nav_r:
    col_info, col_logout = st.columns([2, 1])
    with col_info:
        st.markdown(
            '<div style="text-align:right;padding:8px 0;font-size:11px;color:#9ca3af;">'
            'Logged in as Admin</div>',
            unsafe_allow_html=True,
        )
    with col_logout:
        if st.button("Logout", use_container_width=True):
            st.session_state.admin_auth       = False
            st.session_state.admin_fail_count = 0
            st.rerun()

st.markdown(
    '<hr style="margin:6px 0 16px 0;border:none;border-top:1px solid #e2e8f0;">',
    unsafe_allow_html=True,
)


# ==========================================
# HELPER FUNCTIONS
# ==========================================

def _section(title: str):
    """G3: Bold section heading with divider."""
    st.markdown(
        f'<div style="font-size:15px;font-weight:700;color:#1E2D5E;'
        f'margin:18px 0 6px 0;">{title}</div>',
        unsafe_allow_html=True,
    )

def _divider():
    st.markdown(
        '<hr style="margin:10px 0;border:none;border-top:1px solid #e2e8f0;">',
        unsafe_allow_html=True,
    )


def get_departments() -> list:
    if not os.path.exists(CLIENT_CONTRACTS_DIR):
        return []
    return sorted([
        d for d in os.listdir(CLIENT_CONTRACTS_DIR)
        if os.path.isdir(os.path.join(CLIENT_CONTRACTS_DIR, d))
    ])


def get_dept_files(dept: str) -> list:
    dept_dir = os.path.join(CLIENT_CONTRACTS_DIR, dept)
    files = []
    for ext in ["*.pdf", "*.docx"]:
        for f in glob.glob(os.path.join(dept_dir, ext)):
            stat = os.stat(f)
            files.append({
                "name":       os.path.basename(f),
                "path":       f,
                "size_kb":    round(stat.st_size / 1024, 1),
                "modified":   datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d"),
                "department": dept,
            })
    return sorted(files, key=lambda x: x["name"])


def get_flat_client_files() -> list:
    files = []
    for ext in ["*.pdf", "*.docx"]:
        for f in glob.glob(os.path.join(CLIENT_CONTRACTS_DIR, ext)):
            stat = os.stat(f)
            files.append({
                "name":       os.path.basename(f),
                "path":       f,
                "size_kb":    round(stat.st_size / 1024, 1),
                "modified":   datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d"),
                "department": "(unorganized)",
            })
    return sorted(files, key=lambda x: x["name"])


def get_all_client_files() -> list:
    all_files = get_flat_client_files()
    for dept in get_departments():
        all_files.extend(get_dept_files(dept))
    return sorted(all_files, key=lambda x: (x["department"], x["name"]))


def get_all_client_filenames() -> set:
    return {f["name"] for f in get_all_client_files()}


def find_file_department(filename: str) -> str:
    if os.path.exists(os.path.join(CLIENT_CONTRACTS_DIR, filename)):
        return "clients/ (root)"
    for dept in get_departments():
        if os.path.exists(os.path.join(CLIENT_CONTRACTS_DIR, dept, filename)):
            return dept
    return "unknown"


def get_playbook_files() -> list:
    files = []
    for ext in ["*.pdf", "*.docx"]:
        for f in glob.glob(os.path.join(COMPANY_PLAYBOOK_DIR, ext)):
            stat      = os.stat(f)
            stem      = os.path.splitext(os.path.basename(f))[0]
            safe_stem = "".join(c if c.isalnum() or c in "-_" else "_" for c in stem)
            chroma    = os.path.join(CHROMA_PERSIST_DIR, safe_stem)
            hash_file = os.path.join(chroma, ".playbook_hash")
            built_at  = "—"
            if os.path.exists(hash_file):
                try:
                    data     = json.load(open(hash_file))
                    built_at = data.get("built_at", "—")
                except Exception:
                    pass
            files.append({
                "name":       os.path.basename(f),
                "path":       f,
                "size_kb":    round(stat.st_size / 1024, 1),
                "modified":   datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d"),
                "db_exists":  os.path.exists(chroma) and bool(os.listdir(chroma)),
                "built_at":   built_at,
                "chroma_dir": chroma,
            })
    return sorted(files, key=lambda x: x["name"])


def get_file_risk_status(filename: str) -> str:
    history = logger.get_history_log()
    for entry in reversed(history):
        if entry.get("client_file") == filename:
            if entry.get("high_count",   0) > 0: return "High risk"
            if entry.get("medium_count", 0) > 0: return "Medium risk"
            if entry.get("low_count",    0) > 0: return "Low risk"
            return "Analyzed"
    return "Not analyzed"


def risk_badge_html(status: str) -> str:
    colors_map = {
        "High risk":    ("#fee2e2", "#991b1b"),
        "Medium risk":  ("#fef3c7", "#92400e"),
        "Low risk":     ("#dcfce7", "#166534"),
        "Analyzed":     ("#dcfce7", "#166534"),
        "Not analyzed": ("#f1f5f9", "#64748b"),
    }
    bg, fg = colors_map.get(status, colors_map["Not analyzed"])
    return (
        f'<span style="background:{bg};color:{fg};padding:2px 10px;'
        f'border-radius:12px;font-size:11px;font-weight:600;">{status}</span>'
    )


def filter_entries_by_daterange(entries: list, start: date, end: date) -> list:
    """Filter log entries by a calendar date range (inclusive)."""
    result = []
    for e in entries:
        try:
            entry_date = datetime.fromisoformat(e.get("timestamp", "")).date()
            if start <= entry_date <= end:
                result.append(e)
        except Exception:
            pass
    return result


def save_client_file_with_uniqueness_check(uploaded_file, target_dir: str) -> tuple[bool, str]:
    if not os.path.exists(target_dir):
        os.makedirs(target_dir, exist_ok=True)
    all_existing = get_all_client_filenames()
    if uploaded_file.name in all_existing:
        existing_dept = find_file_department(uploaded_file.name)
        return (False,
                f"**'{uploaded_file.name}'** already exists in **{existing_dept}**. "
                f"Filenames must be unique across all departments. Rename before uploading.")
    dest = os.path.join(target_dir, uploaded_file.name)
    try:
        with open(dest, "wb") as fh:
            fh.write(uploaded_file.getbuffer())
        return (True, f"'{uploaded_file.name}' uploaded successfully.")
    except Exception as e:
        return (False, f"Upload failed: {str(e)}")


# ==========================================
# MAIN TABS
# ==========================================
tab1, tab2, tab3, tab4 = st.tabs([
    "📂 Documents",
    "📊 Token Usage",
    "🕓 History",
    "⬇ Export",
])


# ==========================================
# TAB 1: DOCUMENT MANAGEMENT
# ==========================================
with tab1:

    depts          = get_departments()
    flat_files     = get_flat_client_files()
    pb_files       = get_playbook_files()
    all_dept_files = []
    for d in depts:
        all_dept_files.extend(get_dept_files(d))

    total_contracts = len(all_dept_files) + len(flat_files)
    history         = logger.get_history_log()
    analyzed_names  = {e.get("client_file") for e in history}
    unanalyzed      = sum(1 for f in (all_dept_files + flat_files)
                          if f["name"] not in analyzed_names)

    # D2: Metrics row
    _section("Summary")
    m1, m2, m3 = st.columns(3)
    m1.metric("Departments",     len(depts))
    m2.metric("Total Contracts", total_contracts)
    m3.metric("Unanalyzed",      unanalyzed)

    _divider()

    # D2: Client contracts section
    _section("Client Contracts")

    # Toolbar: department filter | search | status filter
    tf1, tf2, tf3 = st.columns([2, 2, 2])
    with tf1:
        dept_filter_options  = ["All departments"] + depts
        selected_filter_dept = st.selectbox(
            "Department", options=dept_filter_options,
            key="doc_dept_filter", label_visibility="collapsed",
        )
    with tf2:
        search_term = st.text_input(
            "Search", placeholder="Search by filename...",
            key="doc_search", label_visibility="collapsed",
        )
    with tf3:
        # G2: selectbox (non-typeable)
        status_filter = st.selectbox(
            "Status", ["All statuses", "Analyzed", "Unanalyzed"],
            key="doc_status_filter", label_visibility="collapsed",
        )

    # Single row radio for panel toggle — "View" replaces "None"
    panel_mode = st.radio(
        "Panel",
        options=["View", "Upload Contract", "Manage Departments"],
        index=0,
        horizontal=True,
        key="panel_mode_radio",
        label_visibility="collapsed",
    )

    # ── Upload panel ───────────────────────────────────────────────────────────
    if panel_mode == "Upload Contract":
        _divider()
        _section("Upload Contract")
        up_col1, _ = st.columns([3, 1])
        with up_col1:
            # G2: selectbox non-typeable
            upload_dest_dept = st.selectbox(
                "Upload to department",
                options=["(root — no department)"] + depts,
                key="upload_dest_dept",
            )
        new_file = st.file_uploader(
            "Choose file", type=["pdf", "docx"],
            key="tab1_upload", label_visibility="collapsed",
        )
        if new_file:
            target = (
                CLIENT_CONTRACTS_DIR
                if upload_dest_dept == "(root — no department)"
                else os.path.join(CLIENT_CONTRACTS_DIR, upload_dest_dept)
            )
            ok, msg = save_client_file_with_uniqueness_check(new_file, target)
            if ok:
                st.toast(msg, icon="✅")
                st.rerun()
            else:
                st.warning(msg)

    # ── Manage Departments panel ───────────────────────────────────────────────
    if panel_mode == "Manage Departments":
        _divider()
        _section("Create New Department")
        st.caption("All fields are required.")

        cr1, cr2, cr3, cr4, cr5 = st.columns([2, 2, 2, 2, 1])
        with cr1:
            new_dept_name = st.text_input(
                "Dept name", placeholder="Department name *",
                key="new_dept_name_input", label_visibility="collapsed",
            )
        with cr2:
            new_dept_head = st.text_input(
                "Head name", placeholder="Head name *",
                key="new_dept_head_input", label_visibility="collapsed",
            )
        with cr3:
            new_dept_email = st.text_input(
                "Head email", placeholder="Email *",
                key="new_dept_email_input", label_visibility="collapsed",
            )
        with cr4:
            new_dept_phone = st.text_input(
                "Head phone", placeholder="Phone * (10 digits or +country)",
                key="new_dept_phone_input", label_visibility="collapsed",
            )
        with cr5:
            create_clicked = st.button("Create", key="create_dept_btn", use_container_width=True)

        if create_clicked:
            # Fix 5: Run all 4 validators independently — collect ALL errors at once
            errors = []
            ok1, e1 = dept_db.validate_dept_name(new_dept_name.strip())
            ok2, e2 = dept_db.validate_head_name(new_dept_head.strip(), required=True)
            ok3, e3 = dept_db.validate_email(new_dept_email.strip(), required=True)
            ok4, e4 = dept_db.validate_phone(new_dept_phone.strip(), required=True)
            if not ok1: errors.append(e1)
            if not ok2: errors.append(e2)
            if not ok3: errors.append(e3)
            if not ok4: errors.append(e4)

            if errors:
                st.session_state["create_dept_err"] = errors
            else:
                new_path = os.path.join(CLIENT_CONTRACTS_DIR, new_dept_name.strip())
                if os.path.exists(new_path):
                    st.session_state["create_dept_err"] = [f"'{new_dept_name}' already exists."]
                else:
                    st.session_state.pop("create_dept_err", None)
                    os.makedirs(new_path, exist_ok=True)
                    dept_db.upsert_dept(
                        new_dept_name.strip(),
                        new_dept_head.strip(),
                        new_dept_email.strip(),
                        new_dept_phone.strip(),
                    )
                    st.toast(f"Department '{new_dept_name.strip()}' created.", icon="✅")
                    st.rerun()

        # Full-width error display — shows ALL errors
        if st.session_state.get("create_dept_err"):
            err_list = st.session_state["create_dept_err"]
            if isinstance(err_list, list):
                err_msg = "\n\n".join(f"• {e}" for e in err_list)
            else:
                err_msg = str(err_list)
            st.error(err_msg)
            if st.button("Dismiss", key="dismiss_create_err"):
                st.session_state.pop("create_dept_err", None)
                st.rerun()

        # Department heads table
        if depts:
            _divider()
            _section("Department Heads")

            if "editing_dept_head" not in st.session_state:
                st.session_state["editing_dept_head"] = None

            # Table column headers — G3: bold
            th_cols = st.columns([1.5, 1.8, 2, 1.8, 1.5, 1])
            for col, label in zip(th_cols, ["Department", "Head Name", "Email", "Phone", "Updated", "Action"]):
                col.markdown(
                    f'<div style="font-size:11px;font-weight:700;color:#374151;'
                    f'text-transform:uppercase;letter-spacing:0.5px;padding:4px 0;">{label}</div>',
                    unsafe_allow_html=True,
                )
            st.markdown(
                '<hr style="margin:2px 0 6px 0;border:none;border-top:1px solid #e2e8f0;">',
                unsafe_allow_html=True,
            )

            all_db_depts = {d["name"]: d for d in dept_db.get_all_depts()}

            for dept in depts:
                db_row     = all_db_depts.get(dept, {})
                is_editing = st.session_state["editing_dept_head"] == dept
                tr_cols    = st.columns([1.5, 1.8, 2, 1.8, 1.5, 1])

                if is_editing:
                    with tr_cols[0]:
                        edited_dept_name = st.text_input(
                            "Dept", value=dept,
                            key=f"edit_dname_{dept}", label_visibility="collapsed",
                        )
                    with tr_cols[1]:
                        edited_head = st.text_input(
                            "Head", value=db_row.get("head_name", ""),
                            key=f"edit_head_{dept}", label_visibility="collapsed",
                        )
                    with tr_cols[2]:
                        edited_email = st.text_input(
                            "Email", value=db_row.get("head_email", ""),
                            key=f"edit_email_{dept}", label_visibility="collapsed",
                        )
                    with tr_cols[3]:
                        edited_phone = st.text_input(
                            "Phone", value=db_row.get("head_phone", ""),
                            key=f"edit_phone_{dept}", label_visibility="collapsed",
                        )
                    with tr_cols[4]:
                        st.markdown(
                            '<div style="font-size:11px;color:#6b7280;padding:10px 0;">editing...</div>',
                            unsafe_allow_html=True,
                        )
                    with tr_cols[5]:
                        if st.button("Save", key=f"save_{dept}",
                                     use_container_width=True, type="primary"):
                            # D1: all fields required on edit too
                            ok, err = dept_db.validate_all(
                                edited_dept_name.strip(),
                                edited_head.strip(),
                                edited_email.strip(),
                                edited_phone.strip(),
                            )
                            if not ok:
                                st.session_state[f"dept_err_{dept}"] = err
                            else:
                                st.session_state.pop(f"dept_err_{dept}", None)
                                new_dname = edited_dept_name.strip()
                                if new_dname != dept:
                                    old_path = os.path.join(CLIENT_CONTRACTS_DIR, dept)
                                    new_path = os.path.join(CLIENT_CONTRACTS_DIR, new_dname)
                                    if os.path.exists(new_path):
                                        st.session_state[f"dept_err_{dept}"] = f"'{new_dname}' already exists."
                                    else:
                                        try:
                                            os.rename(old_path, new_path)
                                            ok2, err2 = dept_db.rename_dept(dept, new_dname)
                                            if not ok2:
                                                os.rename(new_path, old_path)
                                                st.session_state[f"dept_err_{dept}"] = err2
                                            else:
                                                dept_db.upsert_dept(new_dname, edited_head,
                                                                     edited_email, edited_phone)
                                                st.session_state["editing_dept_head"] = None
                                                st.toast(f"Renamed to '{new_dname}' and saved.", icon="✅")
                                                st.rerun()
                                        except Exception as e:
                                            st.session_state[f"dept_err_{dept}"] = f"Rename failed: {e}"
                                else:
                                    dept_db.upsert_dept(dept, edited_head, edited_email, edited_phone)
                                    st.session_state["editing_dept_head"] = None
                                    st.toast(f"'{dept}' updated.", icon="✅")
                                    st.rerun()

                    # Full-width error outside columns
                    err_key = f"dept_err_{dept}"
                    if st.session_state.get(err_key):
                        st.error(st.session_state[err_key])
                        if st.button("Dismiss", key=f"dismiss_err_{dept}"):
                            st.session_state.pop(err_key, None)
                            st.rerun()
                else:
                    def _cell(val, col, mono=False):
                        style = "font-family:monospace;" if mono else ""
                        col.markdown(
                            f'<div style="font-size:12px;padding:9px 0;{style}">'
                            f'{val if val else "<span style=\'color:#9ca3af;\'>—</span>"}</div>',
                            unsafe_allow_html=True,
                        )

                    _cell(dept,                          tr_cols[0])
                    _cell(db_row.get("head_name",  ""), tr_cols[1])
                    _cell(db_row.get("head_email", ""), tr_cols[2])
                    _cell(db_row.get("head_phone", ""), tr_cols[3])
                    updated = db_row.get("updated_at", "")
                    tr_cols[4].markdown(
                        f'<div style="font-size:11px;color:#6b7280;padding:10px 0;">'
                        f'{updated[:10] if updated else "—"}</div>',
                        unsafe_allow_html=True,
                    )
                    if tr_cols[5].button("Edit", key=f"edit_btn_{dept}", use_container_width=True):
                        st.session_state["editing_dept_head"] = dept
                        st.rerun()

                st.markdown(
                    '<hr style="margin:2px 0;border:none;border-top:0.5px solid #f1f5f9;">',
                    unsafe_allow_html=True,
                )

    # ── Build filtered file list ───────────────────────────────────────────────
    if selected_filter_dept == "All departments":
        display_files = all_dept_files + flat_files
    else:
        display_files = get_dept_files(selected_filter_dept)

    if search_term:
        display_files = [f for f in display_files
                         if search_term.lower() in f["name"].lower()]
    if status_filter == "Analyzed":
        display_files = [f for f in display_files if f["name"] in analyzed_names]
    elif status_filter == "Unanalyzed":
        display_files = [f for f in display_files if f["name"] not in analyzed_names]

    st.markdown(
        f'<div style="font-size:12px;color:#6b7280;margin:8px 0 4px;">'
        f'Showing <strong>{len(display_files)}</strong> file(s)'
        + (f' in <strong>{selected_filter_dept}</strong>'
           if selected_filter_dept != "All departments"
           else ' across all departments')
        + '</div>',
        unsafe_allow_html=True,
    )

    def _file_table_header():
        # 4 cols: File Name | Size | Uploaded | Actions(centered)
        h1, h2, h3, h4 = st.columns([4, 1, 1.5, 3.3])
        for col, label, align in zip(
            [h1, h2, h3, h4],
            ["File Name", "Size", "Uploaded", "Actions"],
            ["left", "left", "left", "center"],
        ):
            col.markdown(
                f'<div style="font-size:11px;font-weight:600;color:#6b7280;'
                f'text-transform:uppercase;letter-spacing:0.5px;padding:5px 0 4px;'
                f'text-align:{align};">{label}</div>',
                unsafe_allow_html=True,
            )
        st.markdown('<hr style="margin:0 0 2px 0;border:none;border-top:1px solid #e2e8f0;">',
                    unsafe_allow_html=True)

    def _file_row(f: dict, dept_name: str):
        # Fix 3: Risk Status column removed — 4 cols matching header
        other_depts = [d for d in depts if d != dept_name]
        if dept_name == "(unorganized)":
            other_depts = depts

        c1, c2, c3, c4 = st.columns([4, 1, 1.5, 3.3])
        with c1:
            st.markdown(
                f'<div style="font-size:13px;padding:9px 0 2px;word-break:break-word;">'
                f'{f["name"]}</div>',
                unsafe_allow_html=True,
            )
        with c2:
            st.markdown(
                f'<div style="font-size:12px;color:#6b7280;padding:10px 0;">'
                f'{f["size_kb"]} KB</div>',
                unsafe_allow_html=True,
            )
        with c3:
            st.markdown(
                f'<div style="font-size:12px;color:#6b7280;padding:10px 0;">'
                f'{f["modified"]}</div>',
                unsafe_allow_html=True,
            )
        with c4:
            act_move, act_del = st.columns([1.4, 1])  # Fix 5: Move to narrower
            with act_move:
                if other_depts:
                    move_to = st.selectbox(
                        "Move",
                        options=["Move to..."] + other_depts,
                        key=f"move_{dept_name}_{f['name']}",
                        label_visibility="collapsed",
                    )
                    if move_to != "Move to...":
                        dest_path = os.path.join(CLIENT_CONTRACTS_DIR, move_to, f["name"])
                        shutil.move(f["path"], dest_path)
                        st.toast(f"Moved '{f['name']}' to {move_to}.", icon="✅")
                        st.rerun()
                else:
                    st.markdown(
                        '<div style="font-size:11px;color:#9ca3af;padding:9px 0;">—</div>',
                        unsafe_allow_html=True,
                    )
            with act_del:
                if st.button("Delete", key=f"del_{dept_name}_{f['name']}",
                             type="secondary", use_container_width=True):
                    fname = f["name"]
                    os.remove(f["path"])
                    st.toast(f"Deleted '{fname}'.", icon="🗑")
                    st.rerun()

        st.markdown('<hr style="margin:2px 0;border:none;border-top:0.5px solid #f1f5f9;">',
                    unsafe_allow_html=True)

    if not display_files:
        st.info("No files match the current filters.")
    elif selected_filter_dept != "All departments":
        _file_table_header()
        for f in display_files:
            _file_row(f, selected_filter_dept)
    else:
        all_db_depts_map = {d["name"]: d for d in dept_db.get_all_depts()}
        groups: dict = {dept: [] for dept in depts}
        groups["(unorganized)"] = []
        for f in display_files:
            dname = f.get("department", "(unorganized)")
            if dname not in groups:
                groups[dname] = []
            groups[dname].append(f)

        for group_dept, group_files in groups.items():
            if not group_files:
                continue
            analyzed_in_group = sum(1 for f in group_files if f["name"] in analyzed_names)
            label    = group_dept if group_dept != "(unorganized)" else "Unorganized Files"
            db_row   = all_db_depts_map.get(group_dept, {})
            head_name  = db_row.get("head_name",  "")
            head_email = db_row.get("head_email", "")
            head_phone = db_row.get("head_phone", "")

            if head_name or head_email or head_phone:
                parts = []
                if head_name:  parts.append(f"<strong>{head_name}</strong>")
                if head_email: parts.append(f'<a href="mailto:{head_email}" style="color:#3b82f6;">{head_email}</a>')
                if head_phone: parts.append(head_phone)
                head_line = '<div style="font-size:11px;color:#6b7280;margin-top:3px;">Head: ' + " &nbsp;·&nbsp; ".join(parts) + "</div>"
            else:
                head_line = '<div style="font-size:11px;color:#9ca3af;font-style:italic;margin-top:3px;">No head assigned</div>'

            st.markdown(
                f'<div style="background:#f8fafc;border-radius:6px;padding:8px 12px;margin:12px 0 6px;">'
                f'<div style="display:flex;align-items:center;gap:10px;">'
                f'<span style="font-size:13px;font-weight:600;">{label}</span>'
                f'<span style="font-size:11px;color:#6b7280;">{len(group_files)} file(s) · {analyzed_in_group} analyzed</span>'
                f'</div>{head_line}</div>',
                unsafe_allow_html=True,
            )
            _file_table_header()
            for f in group_files:
                _file_row(f, group_dept)

    # D2: Playbooks section
    _divider()
    _section("Company Playbooks")
    st.markdown(
        '<div style="font-size:12px;color:#6b7280;margin-bottom:10px;">'
        '<strong>Wipe DB</strong> — deletes only the ChromaDB vector store (file stays). '
        'Rebuild via main app. &nbsp;|&nbsp; '
        '<strong>Delete</strong> — permanently removes the playbook file AND its ChromaDB.'
        '</div>',
        unsafe_allow_html=True,
    )

    if not pb_files:
        st.info("No playbooks found in company_standard/")
    else:
        pbh1, pbh2, pbh3, _ = st.columns([5, 1.4, 1.4, 0.1])
        for col, label, align in zip([pbh1, pbh2, pbh3],
                                     ["File Name · Info", "Wipe DB", "Delete"],
                                     ["left", "center", "center"]):
            col.markdown(
                f'<div style="font-size:11px;font-weight:600;color:#6b7280;'
                f'text-transform:uppercase;letter-spacing:0.5px;padding:5px 0 4px;'
                f'text-align:{align};">{label}</div>',
                unsafe_allow_html=True,
            )
        st.markdown('<hr style="margin:0 0 6px 0;border:none;border-top:1px solid #e2e8f0;">',
                    unsafe_allow_html=True)

        for pb in pb_files:
            db_status  = "ChromaDB built" if pb["db_exists"] else "No ChromaDB"
            db_color   = "#15803d" if pb["db_exists"] else "#dc2626"
            built_lbl  = pb["built_at"][:10] if pb["built_at"] != "—" else "—"

            pb1, pb2, pb3, _ = st.columns([5, 1.4, 1.4, 0.1])
            with pb1:
                st.markdown(
                    f'<div style="padding:8px 0;">'
                    f'<div style="font-size:13px;font-weight:500;margin-bottom:2px;">{pb["name"]}</div>'
                    f'<div style="font-size:11px;color:#6b7280;">'
                    f'{pb["size_kb"]} KB &nbsp;·&nbsp; '
                    f'<span style="color:{db_color};">{db_status}</span>'
                    f' &nbsp;·&nbsp; built {built_lbl}</div></div>',
                    unsafe_allow_html=True,
                )
            with pb2:
                if pb["db_exists"]:
                    if st.button("Wipe DB", key=f"wipe_{pb['name']}", use_container_width=True):
                        shutil.rmtree(pb["chroma_dir"])
                        st.toast(f"ChromaDB wiped for '{pb['name']}'.", icon="🗑")
                        st.rerun()
                else:
                    st.markdown('<div style="font-size:11px;color:#9ca3af;padding:10px 0;text-align:center;">No DB</div>',
                                unsafe_allow_html=True)
            with pb3:
                if st.button("Delete", key=f"del_pb_{pb['name']}",
                             type="secondary", use_container_width=True):
                    os.remove(pb["path"])
                    if pb["db_exists"]:
                        shutil.rmtree(pb["chroma_dir"])
                    st.toast(f"Deleted '{pb['name']}'.", icon="🗑")
                    st.rerun()

            st.markdown('<hr style="margin:4px 0;border:none;border-top:0.5px solid #f1f5f9;">',
                        unsafe_allow_html=True)

    st.caption("Upload a new playbook to company_standard/")
    new_pb = st.file_uploader("Upload new playbook", type=["pdf", "docx"],
                               key="upload_new_pb", label_visibility="collapsed")
    if new_pb:
        dest = os.path.join(COMPANY_PLAYBOOK_DIR, new_pb.name)
        if os.path.exists(dest):
            st.warning(f"'{new_pb.name}' already exists.")
        else:
            os.makedirs(COMPANY_PLAYBOOK_DIR, exist_ok=True)
            with open(dest, "wb") as fh:
                fh.write(new_pb.getbuffer())
            st.toast(f"Uploaded '{new_pb.name}'. Rebuild ChromaDB from main app.", icon="✅")
            st.rerun()


# ==========================================
# TAB 2: TOKEN USAGE ANALYTICS
# ==========================================
with tab2:

    _section("Date Range")
    # T2: calendar date picker — min_value prevents clearing by backspace
    today      = date.today()
    default_start = today - timedelta(days=30)
    date_range = st.date_input(
        "Select date range",
        value=(default_start, today),
        min_value=date(2020, 1, 1),
        max_value=today,
        key="token_date_range",
        label_visibility="collapsed",
    )
    # Handle single date vs range
    if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
        start_date, end_date = date_range[0], date_range[1]
    else:
        start_date = end_date = date_range if not isinstance(date_range, (list, tuple)) else date_range[0]

    all_token_entries = logger.get_token_log()
    token_entries     = filter_entries_by_daterange(all_token_entries, start_date, end_date)

    _divider()

    if not token_entries:
        st.info("No token usage data in the selected date range. Run the pipeline to start logging.")
    else:
        _section("Summary Metrics")
        total_tokens = sum(e.get("total_tokens", 0) for e in token_entries)
        total_cost   = sum(e.get("cost_usd",     0) for e in token_entries)
        total_calls  = len(token_entries)
        avg_dur      = sum(e.get("duration_ms", 0) for e in token_entries) // max(len(token_entries), 1)

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Tokens",   f"{total_tokens:,}")
        m2.metric("Estimated Cost", f"${total_cost:.4f}")
        m3.metric("API Calls",      total_calls)
        m4.metric("Avg Duration",   f"{avg_dur:,} ms")

        _divider()

        # T1: Tool breakdown chart with type toggle + expand
        _section("Token Usage by Tool")
        import pandas as pd

        tool_data: dict = {}
        for e in token_entries:
            t = e.get("tool_name", "Unknown")
            if t not in tool_data:
                tool_data[t] = {"calls": 0, "total_tokens": 0, "cost_usd": 0.0}
            tool_data[t]["calls"]        += 1
            tool_data[t]["total_tokens"] += e.get("total_tokens", 0)
            tool_data[t]["cost_usd"]     += e.get("cost_usd", 0.0)

        if tool_data:
            ctrl1, ctrl2, ctrl3 = st.columns([2, 2, 4])
            with ctrl1:
                # G2: selectbox non-typeable
                chart_type = st.selectbox(
                    "Chart type",
                    ["Bar", "Line"],
                    key="tool_chart_type",
                )
            with ctrl2:
                expand_chart = st.checkbox("Expand chart", key="expand_tool_chart")

            chart_height = 450 if expand_chart else 260

            df_chart = pd.DataFrame({
                "Tool": list(tool_data.keys()),
                "Tokens": [v["total_tokens"] for v in tool_data.values()],
            }).set_index("Tool")

            if chart_type == "Bar":
                st.bar_chart(df_chart, height=chart_height)
            else:
                st.line_chart(df_chart, height=chart_height)

        # Tool breakdown table
        col_headers = ["Tool", "Calls", "Total Tokens", "Input Tokens", "Output Tokens", "Cost (USD)"]
        tool_rows   = []
        for tool, vals in tool_data.items():
            tin  = sum(e.get("input_tokens",  0) for e in token_entries if e.get("tool_name") == tool)
            tout = sum(e.get("output_tokens", 0) for e in token_entries if e.get("tool_name") == tool)
            tool_rows.append([tool, vals["calls"],
                               f"{vals['total_tokens']:,}", f"{tin:,}", f"{tout:,}",
                               f"${vals['cost_usd']:.6f}"])
        if tool_rows:
            st.dataframe(pd.DataFrame(tool_rows, columns=col_headers),
                         use_container_width=True, hide_index=True)

        _divider()

        # T1: Daily chart with toggle
        _section("Daily Token Usage")
        daily: dict = {}
        for e in token_entries:
            try:
                day = e["timestamp"][:10]
                daily[day] = daily.get(day, 0) + e.get("total_tokens", 0)
            except Exception:
                pass
        if daily:
            ctrl_d1, ctrl_d2, _ = st.columns([2, 2, 4])
            with ctrl_d1:
                daily_chart_type = st.selectbox(
                    "Chart type", ["Bar", "Line"], key="daily_chart_type"
                )
            with ctrl_d2:
                expand_daily = st.checkbox("Expand chart", key="expand_daily_chart")

            daily_height = 450 if expand_daily else 260
            df_daily = pd.DataFrame(
                sorted(daily.items()), columns=["Date", "Tokens"]
            ).set_index("Date")
            if daily_chart_type == "Bar":
                st.bar_chart(df_daily, height=daily_height)
            else:
                st.line_chart(df_daily, height=daily_height)

        _divider()

        _section("Raw Token Log")
        # G2: selectbox non-typeable
        tool_filter = st.selectbox(
            "Filter by tool",
            ["All tools"] + sorted({e.get("tool_name", "") for e in token_entries}),
            key="token_tool_filter",
        )
        display_entries = [
            e for e in reversed(token_entries)
            if tool_filter == "All tools" or e.get("tool_name") == tool_filter
        ]
        rows = []
        for e in display_entries[:100]:
            rows.append({
                "Timestamp":     e.get("timestamp", ""),
                "Tool":          e.get("tool_name", ""),
                "Input Tokens":  e.get("input_tokens",  0),
                "Output Tokens": e.get("output_tokens", 0),
                "Total Tokens":  e.get("total_tokens",  0),
                "Cost (USD)":    f"${e.get('cost_usd', 0):.6f}",
                "Duration (ms)": e.get("duration_ms",  0),
                "Client File":   e.get("client_file",  ""),
            })
        if rows:
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        st.markdown("")
        if st.button("Clear Token Log", type="secondary"):
            logger.clear_token_log()
            st.toast("Token log cleared.", icon="🗑")
            st.rerun()


# ==========================================
# TAB 3: HISTORY LOG  (H1: simplified)
# ==========================================
with tab3:

    history_entries = logger.get_history_log()

    if not history_entries:
        st.info("No analysis history yet. Run the full pipeline to start logging.")
    else:
        hist_summary = logger.get_history_summary()

        _section("Summary")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Runs",        hist_summary["total_runs"])
        m2.metric("High Risk Runs",    hist_summary["high_risk_runs"])
        m3.metric("Avg Clauses / Run", hist_summary["avg_clauses"])
        m4.metric("Top Department",    hist_summary["top_department"])

        _divider()
        _section("Filter")

        # H1: only filename search + department filter (risk level filter removed)
        col_s, col_d = st.columns(2)
        with col_s:
            search_hist = st.text_input(
                "Search", placeholder="Filter by filename...",
                key="hist_search", label_visibility="collapsed",
            )
        with col_d:
            dept_opts = ["All departments"] + sorted({
                e.get("department", "") or "Unknown" for e in history_entries
            })
            # G2: selectbox non-typeable
            dept_filter = st.selectbox(
                "Department", dept_opts,
                key="hist_dept_filter", label_visibility="collapsed",
            )

        filtered_hist = list(reversed(history_entries))
        if search_hist:
            filtered_hist = [e for e in filtered_hist
                             if search_hist.lower() in e.get("client_file", "").lower()]
        if dept_filter != "All departments":
            filtered_hist = [e for e in filtered_hist
                             if (e.get("department") or "Unknown") == dept_filter]

        _divider()
        st.markdown(f"Showing **{len(filtered_hist)}** run(s)")

        for entry in filtered_hist:
            h = entry.get("high_count",   0)
            m = entry.get("medium_count", 0)
            l = entry.get("low_count",    0)

            risk_pills = ""
            for count, label_r, bg, fg in [
                (h, "High",   "#fee2e2", "#991b1b"),
                (m, "Medium", "#fef3c7", "#92400e"),
                (l, "Low",    "#dcfce7", "#166534"),
            ]:
                if count:
                    risk_pills += (
                        f'<span style="background:{bg};color:{fg};padding:2px 8px;'
                        f'border-radius:10px;font-size:11px;margin-right:4px;">'
                        f'{count} {label_r}</span>'
                    )
            if not (h or m or l):
                risk_pills = (
                    '<span style="background:#f1f5f9;color:#6b7280;padding:2px 8px;'
                    'border-radius:10px;font-size:11px;">No risk data</span>'
                )

            dept_label = entry.get("department") or "No department"

            with st.expander(
                f"{entry.get('client_file', 'Unknown')}  ·  "
                f"{entry.get('timestamp', '')[:16]}",
                expanded=False,
            ):
                st.markdown(
                    f'<div style="margin-bottom:10px;">{risk_pills}</div>',
                    unsafe_allow_html=True,
                )
                col_a, col_b, col_c = st.columns(3)
                col_a.markdown(f"**Department:** {dept_label}")
                col_b.markdown(f"**Playbook:** {entry.get('playbook_file', '—')}")
                col_c.markdown(f"**Total Tokens:** {entry.get('total_tokens', 0):,}")

                clauses = entry.get("clauses_selected", [])
                if clauses:
                    st.markdown(f"**Clauses analyzed:** {', '.join(clauses)}")

                clause_detail = entry.get("clause_detail", [])
                if clause_detail:
                    st.markdown("**Clause-level results:**")
                    for cd in clause_detail:
                        level = cd.get("risk_level", "—")
                        name  = cd.get("clause_name", "—")
                        conf  = cd.get("factual_conflict", "—")
                        color = {
                            "High":   "#dc2626",
                            "Medium": "#d97706",
                            "Low":    "#16a34a",
                        }.get(level, "#6b7280")
                        st.markdown(
                            f'<div style="border-left:3px solid {color};padding:6px 10px;'
                            f'margin:4px 0;background:#f8fafc;border-radius:0 6px 6px 0;">'
                            f'<strong>{name}</strong> — <span style="color:{color};">{level}</span><br>'
                            f'<span style="font-size:12px;color:#6b7280;">{conf}</span>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                if st.button("Delete this run",
                             key=f"del_run_{entry.get('run_id', '')}",
                             type="secondary"):
                    logger.delete_history_entry(entry.get("run_id", ""))
                    st.toast("Run deleted.", icon="🗑")
                    st.rerun()


# ==========================================
# TAB 4: EXPORT
# ==========================================
with tab4:

    _section("Generate Admin Report")
    st.markdown(
        "Download a PDF or Excel report with full token usage, "
        "department file listing, clause-level risk details, and raw token log."
    )

    _divider()
    _section("Report Date Range")

    # E1: calendar date picker — min_value prevents clearing by backspace
    today_e = date.today()
    default_start_e = today_e - timedelta(days=30)
    export_date_range = st.date_input(
        "Select date range for report",
        value=(default_start_e, today_e),
        min_value=date(2020, 1, 1),
        max_value=today_e,
        key="export_date_range",
        label_visibility="collapsed",
    )
    if isinstance(export_date_range, (list, tuple)) and len(export_date_range) == 2:
        exp_start, exp_end = export_date_range[0], export_date_range[1]
    else:
        exp_start = exp_end = (
            export_date_range
            if not isinstance(export_date_range, (list, tuple))
            else export_date_range[0]
        )

    _divider()
    col_pdf, col_xlsx = st.columns(2)

    # ── PDF color constants ────────────────────────────────────────────────────
    DARK_BLUE  = rl_colors.HexColor("#1E2D5E")
    LIGHT_BLUE = rl_colors.HexColor("#E6F1FB")
    ROW_ALT    = rl_colors.HexColor("#f8fafc")
    RED_BG     = rl_colors.HexColor("#fee2e2")
    ORANGE_BG  = rl_colors.HexColor("#fef3c7")
    GREEN_BG   = rl_colors.HexColor("#dcfce7")

    def _tbl_style(extra=None):
        base = [
            ("BACKGROUND",     (0, 0), (-1, 0),  DARK_BLUE),
            ("TEXTCOLOR",      (0, 0), (-1, 0),  rl_colors.white),
            ("FONTNAME",       (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",       (0, 0), (-1, 0),  9),
            ("FONTSIZE",       (0, 1), (-1, -1), 8),
            ("WORDWRAP",       (0, 0), (-1, -1), True),
            ("VALIGN",         (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, ROW_ALT]),
            ("GRID",           (0, 0), (-1, -1), 0.4, rl_colors.lightgrey),
            ("TOPPADDING",     (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 5),
            ("LEFTPADDING",    (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 6),
        ]
        if extra:
            base.extend(extra)
        return TableStyle(base)

    # ── PDF ───────────────────────────────────────────────────────────────────
    with col_pdf:
        _section("PDF Report")
        st.caption("Overview metrics · Token breakdown · Department files · History · Raw log")

        if st.button("Generate PDF", use_container_width=True):
            with st.spinner("Building PDF..."):
                try:
                    token_data   = filter_entries_by_daterange(
                        logger.get_token_log(), exp_start, exp_end)
                    history_data = filter_entries_by_daterange(
                        logger.get_history_log(), exp_start, exp_end)

                    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

                    buffer = BytesIO()
                    doc = SimpleDocTemplate(
                        buffer, pagesize=LETTER,
                        rightMargin=36, leftMargin=36,
                        topMargin=44, bottomMargin=36,
                    )
                    PAGE_W   = 540
                    styles   = getSampleStyleSheet()
                    n_style  = styles["Normal"]
                    sm_style = ParagraphStyle("sm", parent=n_style, fontSize=8, leading=11)
                    cell_style = ParagraphStyle("cell", parent=n_style, fontSize=8,
                                               leading=11, wordWrap="LTR")
                    elements = []

                    elements.append(Paragraph("Admin Report — Legal Contract Analyzer", styles["Title"]))
                    elements.append(Paragraph(
                        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  ·  "
                        f"Period: {exp_start} to {exp_end}  ·  "
                        f"Departments: {len(get_departments())}",
                        sm_style,
                    ))
                    elements.append(Spacer(1, 12))
                    elements.append(HRFlowable(width="100%", thickness=2, color=DARK_BLUE))
                    elements.append(Spacer(1, 16))

                    # Section 1: Overview
                    elements.append(Paragraph("1. Overview Metrics", styles["Heading2"]))
                    elements.append(Spacer(1, 6))
                    all_files_pdf = get_all_client_files()
                    total_t    = sum(e.get("total_tokens", 0) for e in token_data)
                    total_cost_v = sum(e.get("cost_usd",   0) for e in token_data)

                    ov_data = [["Metric", "Value"],
                               ["Total API calls",        str(len(token_data))],
                               ["Total tokens used",      f"{total_t:,}"],
                               ["Estimated cost (USD)",   f"${total_cost_v:.6f}"],
                               ["Pipeline runs",          str(len(history_data))],
                               ["Runs with HIGH risk",    str(sum(1 for e in history_data if e.get("high_count", 0) > 0))],
                               ["Departments",            str(len(get_departments()))],
                               ["Total client contracts", str(len(all_files_pdf))]]
                    ov_tbl = Table(ov_data, colWidths=[300, 240])
                    ov_tbl.setStyle(_tbl_style())
                    elements.append(ov_tbl)
                    elements.append(Spacer(1, 18))
                    elements.append(HRFlowable(width="100%", thickness=0.5, color=rl_colors.lightgrey))
                    elements.append(Spacer(1, 14))

                    # Section 2: Token by tool
                    elements.append(Paragraph("2. Token Usage by Tool", styles["Heading2"]))
                    elements.append(Spacer(1, 6))
                    tool_sum: dict = {}
                    for e in token_data:
                        t = e.get("tool_name", "Unknown")
                        if t not in tool_sum:
                            tool_sum[t] = {"calls": 0, "inp": 0, "out": 0,
                                           "tot": 0, "cost": 0.0, "dur": 0}
                        tool_sum[t]["calls"] += 1
                        tool_sum[t]["inp"]   += e.get("input_tokens",  0)
                        tool_sum[t]["out"]   += e.get("output_tokens", 0)
                        tool_sum[t]["tot"]   += e.get("total_tokens",  0)
                        tool_sum[t]["cost"]  += e.get("cost_usd",      0.0)
                        tool_sum[t]["dur"]   += e.get("duration_ms",   0)

                    if tool_sum:
                        td = [["Tool", "Calls", "Input", "Output", "Total", "Cost (USD)", "Avg ms"]]
                        for tool, v in tool_sum.items():
                            avg_d = v["dur"] // max(v["calls"], 1)
                            td.append([tool, str(v["calls"]),
                                       f"{v['inp']:,}", f"{v['out']:,}",
                                       f"{v['tot']:,}", f"${v['cost']:.6f}", f"{avg_d:,}"])
                        td.append([
                            "TOTAL", str(len(token_data)),
                            f"{sum(v['inp']  for v in tool_sum.values()):,}",
                            f"{sum(v['out']  for v in tool_sum.values()):,}",
                            f"{total_t:,}", f"${total_cost_v:.6f}", "—",
                        ])
                        tl_tbl = Table(td, colWidths=[130, 38, 72, 72, 72, 86, 70])
                        tl_tbl.setStyle(_tbl_style([
                            ("FONTNAME",   (0, -1), (-1, -1), "Helvetica-Bold"),
                            ("BACKGROUND", (0, -1), (-1, -1), LIGHT_BLUE),
                        ]))
                        elements.append(tl_tbl)
                    else:
                        elements.append(Paragraph("No token data in selected period.", n_style))

                    elements.append(Spacer(1, 18))
                    elements.append(HRFlowable(width="100%", thickness=0.5, color=rl_colors.lightgrey))
                    elements.append(Spacer(1, 14))

                    # Section 3: Dept file listing
                    elements.append(Paragraph("3. Department & File Listing", styles["Heading2"]))
                    elements.append(Spacer(1, 6))
                    for dept in get_departments():
                        head = (dept_db.get_dept(dept) or {}).get("head_name", "")
                        head_info = f" · Head: {head}" if head else ""
                        elements.append(Paragraph(f"Department: {dept}{head_info}", styles["Heading3"]))
                        dfiles = get_dept_files(dept)
                        if not dfiles:
                            elements.append(Paragraph("  No files.", sm_style))
                        else:
                            df_data = [["File Name", "Size (KB)", "Modified", "Risk Status"]]
                            for f in dfiles:
                                risk = get_file_risk_status(f["name"])
                                df_data.append([Paragraph(f["name"], cell_style),
                                               str(f["size_kb"]), f["modified"], risk])
                            extra = []
                            for ri, row in enumerate(df_data[1:], 1):
                                r = row[3]
                                if r == "High risk":   extra.append(("BACKGROUND", (3, ri), (3, ri), RED_BG))
                                elif r == "Medium risk": extra.append(("BACKGROUND", (3, ri), (3, ri), ORANGE_BG))
                                elif r in ("Low risk", "Analyzed"): extra.append(("BACKGROUND", (3, ri), (3, ri), GREEN_BG))
                            df_tbl = Table(df_data, colWidths=[250, 60, 90, 140])
                            df_tbl.setStyle(_tbl_style(extra))
                            elements.append(df_tbl)
                        elements.append(Spacer(1, 10))

                    flat = get_flat_client_files()
                    if flat:
                        elements.append(Paragraph("Unorganized files", styles["Heading3"]))
                        fl_data = [["File Name", "Size (KB)", "Modified", "Risk Status"]]
                        for f in flat:
                            fl_data.append([Paragraph(f["name"], cell_style),
                                           str(f["size_kb"]), f["modified"],
                                           get_file_risk_status(f["name"])])
                        fl_tbl = Table(fl_data, colWidths=[250, 60, 90, 140])
                        fl_tbl.setStyle(_tbl_style())
                        elements.append(fl_tbl)

                    elements.append(Spacer(1, 18))
                    elements.append(HRFlowable(width="100%", thickness=0.5, color=rl_colors.lightgrey))
                    elements.append(Spacer(1, 14))

                    # Section 4: History
                    elements.append(Paragraph("4. Analysis History", styles["Heading2"]))
                    elements.append(Spacer(1, 6))
                    if not history_data:
                        elements.append(Paragraph("No history in selected period.", n_style))
                    else:
                        for entry in reversed(history_data):
                            h_c = entry.get("high_count",   0)
                            m_c = entry.get("medium_count", 0)
                            l_c = entry.get("low_count",    0)
                            dept_h = (dept_db.get_dept(entry.get("department") or "") or {}).get("head_name", "")
                            dept_display = entry.get("department") or "—"
                            if dept_h:
                                dept_display += f" (Head: {dept_h})"
                            elements.append(Paragraph(
                                f"<b>{entry.get('client_file', '—')}</b>  "
                                f"| {entry.get('timestamp', '')[:16]}  "
                                f"| Dept: {dept_display}  "
                                f"| Tokens: {entry.get('total_tokens', 0):,}",
                                ParagraphStyle("run_hdr", parent=n_style,
                                               fontSize=8, fontName="Helvetica-Bold"),
                            ))
                            elements.append(Spacer(1, 4))

                            rs_data = [["Clauses", "High", "Medium", "Low"],
                                       [str(entry.get("clause_count", 0)), str(h_c), str(m_c), str(l_c)]]
                            rs_extra = []
                            if h_c: rs_extra.append(("BACKGROUND", (1, 1), (1, 1), RED_BG))
                            if m_c: rs_extra.append(("BACKGROUND", (2, 1), (2, 1), ORANGE_BG))
                            if l_c: rs_extra.append(("BACKGROUND", (3, 1), (3, 1), GREEN_BG))
                            rs_tbl = Table(rs_data, colWidths=[135, 135, 135, 135])
                            rs_tbl.setStyle(_tbl_style(rs_extra))
                            elements.append(rs_tbl)
                            elements.append(Spacer(1, 4))

                            clause_detail = entry.get("clause_detail", [])
                            if clause_detail:
                                cd_data = [["Clause Name", "Risk", "Factual Conflict"]]
                                for cd in clause_detail:
                                    level = cd.get("risk_level", "—")
                                    cd_data.append([
                                        Paragraph(cd.get("clause_name", "—"), cell_style),
                                        level,
                                        Paragraph(cd.get("factual_conflict", "—"), cell_style),
                                    ])
                                cd_extra = []
                                for ri, row in enumerate(cd_data[1:], 1):
                                    lvl = row[1]
                                    if lvl == "High":   cd_extra.append(("BACKGROUND", (1, ri), (1, ri), RED_BG))
                                    elif lvl == "Medium": cd_extra.append(("BACKGROUND", (1, ri), (1, ri), ORANGE_BG))
                                    elif lvl == "Low": cd_extra.append(("BACKGROUND", (1, ri), (1, ri), GREEN_BG))
                                cd_tbl = Table(cd_data, colWidths=[160, 60, 320])
                                cd_tbl.setStyle(_tbl_style(cd_extra))
                                elements.append(cd_tbl)
                            elements.append(Spacer(1, 12))

                    elements.append(HRFlowable(width="100%", thickness=0.5, color=rl_colors.lightgrey))
                    elements.append(Spacer(1, 14))

                    # Section 5: Raw token log
                    elements.append(Paragraph("5. Raw Token Log", styles["Heading2"]))
                    elements.append(Spacer(1, 6))
                    if not token_data:
                        elements.append(Paragraph("No token data.", n_style))
                    else:
                        rl_data = [["Timestamp", "Tool", "In", "Out", "Total", "Cost", "ms", "Client"]]
                        for e in reversed(token_data):
                            cf = e.get("client_file", "")
                            cf_short = (cf[:28] + "…") if len(cf) > 28 else cf
                            rl_data.append([
                                e.get("timestamp", "")[:16],
                                e.get("tool_name",     "")[:20],
                                str(e.get("input_tokens",  0)),
                                str(e.get("output_tokens", 0)),
                                str(e.get("total_tokens",  0)),
                                f"${e.get('cost_usd', 0):.6f}",
                                str(e.get("duration_ms",   0)),
                                cf_short,
                            ])
                        rl_tbl = Table(rl_data, colWidths=[100, 105, 35, 40, 45, 72, 40, 103])
                        rl_tbl.setStyle(_tbl_style())
                        elements.append(rl_tbl)

                    doc.build(elements)
                    buffer.seek(0)

                    st.download_button(
                        label="Download PDF",
                        data=buffer,
                        file_name=f"Admin_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        type="primary",
                    )
                except Exception as ex:
                    st.error(f"PDF generation failed: {ex}")

    # ── Excel ─────────────────────────────────────────────────────────────────
    with col_xlsx:
        _section("Excel Report")
        st.caption("5 sheets: Overview · Token Breakdown · Dept Files · History · Raw Log")

        if st.button("Generate Excel", use_container_width=True):
            with st.spinner("Building Excel..."):
                try:
                    token_data   = filter_entries_by_daterange(
                        logger.get_token_log(), exp_start, exp_end)
                    history_data = filter_entries_by_daterange(
                        logger.get_history_log(), exp_start, exp_end)

                    wb = openpyxl.Workbook()

                    hdr_font    = Font(bold=True, color="FFFFFF")
                    hdr_fill    = PatternFill("solid", fgColor="1E2D5E")
                    total_fill  = PatternFill("solid", fgColor="DBEAFE")
                    red_fill    = PatternFill("solid", fgColor="FEE2E2")
                    orange_fill = PatternFill("solid", fgColor="FEF3C7")
                    green_fill  = PatternFill("solid", fgColor="DCFCE7")
                    grey_fill   = PatternFill("solid", fgColor="F1F5F9")
                    wrap_al     = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    thin_bd     = Border(
                        left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"),  bottom=Side(style="thin"),
                    )

                    def _style_hdr(ws, row=1):
                        for cell in ws[row]:
                            cell.font      = hdr_font
                            cell.fill      = hdr_fill
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.border    = thin_bd
                        ws.row_dimensions[row].height = 20

                    def _auto_w(ws, max_w=55):
                        for col in ws.columns:
                            mx = max((len(str(c.value or "")) for c in col), default=8)
                            ws.column_dimensions[
                                get_column_letter(col[0].column)
                            ].width = min(mx + 3, max_w)

                    def _add_border(ws):
                        for row in ws.iter_rows():
                            for cell in row:
                                if cell.value is not None:
                                    cell.border = thin_bd

                    all_analyzed_xl = {e.get("client_file") for e in logger.get_history_log()}
                    all_files_xl    = get_all_client_files()
                    dept_meta_xl    = {d["name"]: d for d in dept_db.get_all_depts()}

                    # Sheet 1: Overview
                    ws1       = wb.active
                    ws1.title = "Overview"
                    total_t_xl    = sum(e.get("total_tokens", 0) for e in token_data)
                    total_cost_xl = sum(e.get("cost_usd",     0) for e in token_data)
                    ws1.append(["Metric", "Value"])
                    _style_hdr(ws1)
                    for lbl, val in [
                        ("Report period",          f"{exp_start} to {exp_end}"),
                        ("Generated at",            datetime.now().strftime("%Y-%m-%d %H:%M")),
                        ("Total API calls",         len(token_data)),
                        ("Total tokens used",       total_t_xl),
                        ("Estimated cost (USD)",    f"${total_cost_xl:.6f}"),
                        ("Pipeline runs",           len(history_data)),
                        ("Runs with HIGH risk",     sum(1 for e in history_data if e.get("high_count", 0) > 0)),
                        ("Total departments",       len(get_departments())),
                        ("Total client contracts",  len(all_files_xl)),
                    ]:
                        ws1.append([lbl, val])
                    _auto_w(ws1)
                    _add_border(ws1)

                    # Sheet 2: Token breakdown
                    ws2 = wb.create_sheet("Token Breakdown")
                    ws2.append(["Tool", "Calls", "Input", "Output", "Total", "Cost (USD)", "Avg ms"])
                    _style_hdr(ws2)
                    tool_xl: dict = {}
                    for e in token_data:
                        t = e.get("tool_name", "Unknown")
                        if t not in tool_xl:
                            tool_xl[t] = {"calls": 0, "inp": 0, "out": 0,
                                          "tot": 0, "cost": 0.0, "dur": 0}
                        tool_xl[t]["calls"] += 1
                        tool_xl[t]["inp"]   += e.get("input_tokens",  0)
                        tool_xl[t]["out"]   += e.get("output_tokens", 0)
                        tool_xl[t]["tot"]   += e.get("total_tokens",  0)
                        tool_xl[t]["cost"]  += e.get("cost_usd",      0.0)
                        tool_xl[t]["dur"]   += e.get("duration_ms",   0)
                    for tool, v in tool_xl.items():
                        ws2.append([tool, v["calls"], v["inp"], v["out"],
                                    v["tot"], round(v["cost"], 6), v["dur"] // max(v["calls"], 1)])
                    ws2.append(["TOTAL",
                                sum(v["calls"] for v in tool_xl.values()),
                                sum(v["inp"]   for v in tool_xl.values()),
                                sum(v["out"]   for v in tool_xl.values()),
                                total_t_xl, round(total_cost_xl, 6), "—"])
                    for cell in ws2[ws2.max_row]:
                        cell.fill = total_fill
                        cell.font = Font(bold=True)
                    _auto_w(ws2)
                    _add_border(ws2)

                    # Sheet 3: Department files
                    ws3 = wb.create_sheet("Department Files")
                    ws3.append(["Department", "Head", "Email", "Phone",
                                "File Name", "Size (KB)", "Modified", "Risk Status", "Analyzed"])
                    _style_hdr(ws3)
                    for dept in get_departments():
                        db_row_xl = dept_meta_xl.get(dept, {})
                        head  = db_row_xl.get("head_name",  "")
                        email = db_row_xl.get("head_email", "")
                        phone = db_row_xl.get("head_phone", "")
                        dfiles = get_dept_files(dept)
                        if not dfiles:
                            ws3.append([dept, head, email, phone, "(no files)", "", "", "", ""])
                            ws3.cell(ws3.max_row, 1).fill = grey_fill
                        else:
                            for f in dfiles:
                                risk     = get_file_risk_status(f["name"])
                                analyzed = "Yes" if f["name"] in all_analyzed_xl else "No"
                                ws3.append([dept, head, email, phone,
                                           f["name"], f["size_kb"], f["modified"], risk, analyzed])
                                lr = ws3.max_row
                                if risk == "High risk":   ws3.cell(lr, 8).fill = red_fill
                                elif risk == "Medium risk": ws3.cell(lr, 8).fill = orange_fill
                                elif risk in ("Low risk", "Analyzed"): ws3.cell(lr, 8).fill = green_fill
                    for f in get_flat_client_files():
                        risk     = get_file_risk_status(f["name"])
                        analyzed = "Yes" if f["name"] in all_analyzed_xl else "No"
                        ws3.append(["(unorganized)", "", "", "",
                                   f["name"], f["size_kb"], f["modified"], risk, analyzed])
                    _auto_w(ws3)
                    _add_border(ws3)

                    # Sheet 4: Analysis history
                    ws4 = wb.create_sheet("Analysis History")
                    ws4.append(["Timestamp", "Client File", "Department", "Dept Head",
                                "Playbook", "Clauses", "High", "Medium", "Low",
                                "Tokens", "Clause Name", "Risk Level", "Factual Conflict"])
                    _style_hdr(ws4)
                    for e in reversed(history_data):
                        h_v  = e.get("high_count",   0)
                        m_v  = e.get("medium_count", 0)
                        l_v  = e.get("low_count",    0)
                        dept = e.get("department", "") or "—"
                        head = dept_meta_xl.get(dept, {}).get("head_name", "") if dept != "—" else ""
                        for cd in (e.get("clause_detail") or [{"clause_name": "—", "risk_level": "—", "factual_conflict": "—"}]):
                            ws4.append([
                                e.get("timestamp",    ""),
                                e.get("client_file",  ""),
                                dept, head,
                                e.get("playbook_file",""),
                                e.get("clause_count", 0),
                                h_v, m_v, l_v,
                                e.get("total_tokens", 0),
                                cd.get("clause_name",      ""),
                                cd.get("risk_level",       ""),
                                cd.get("factual_conflict", ""),
                            ])
                            lr  = ws4.max_row
                            lvl = cd.get("risk_level", "")
                            if h_v and lvl == "High":   ws4.cell(lr, 12).fill = red_fill
                            elif m_v and lvl == "Medium": ws4.cell(lr, 12).fill = orange_fill
                            elif lvl == "Low":           ws4.cell(lr, 12).fill = green_fill
                            ws4.cell(lr, 13).alignment = wrap_al
                    _auto_w(ws4)
                    _add_border(ws4)

                    # Sheet 5: Raw token log
                    ws5 = wb.create_sheet("Token Log")
                    ws5.append(["Timestamp", "Tool", "Input", "Output", "Total",
                                "Cost (USD)", "Duration (ms)", "Client File", "Department"])
                    _style_hdr(ws5)
                    for e in reversed(token_data):
                        ws5.append([
                            e.get("timestamp",     ""),
                            e.get("tool_name",     ""),
                            e.get("input_tokens",  0),
                            e.get("output_tokens", 0),
                            e.get("total_tokens",  0),
                            round(e.get("cost_usd", 0.0), 6),
                            e.get("duration_ms",   0),
                            e.get("client_file",   ""),
                            e.get("department",    ""),
                        ])
                    _auto_w(ws5)
                    _add_border(ws5)

                    xlsx_buf = BytesIO()
                    wb.save(xlsx_buf)
                    xlsx_buf.seek(0)

                    st.download_button(
                        label="Download Excel",
                        data=xlsx_buf,
                        file_name=f"Admin_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                    )
                except Exception as ex:
                    st.error(f"Excel generation failed: {ex}")